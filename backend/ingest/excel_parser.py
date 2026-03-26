from __future__ import annotations

import io
from typing import BinaryIO

import igraph as ig

from .csv_parser import parse_csv_auto


def parse_excel(file: BinaryIO, directed: bool = False) -> ig.Graph:
    """
    Parse an Excel (.xlsx) file as a graph.

    Reads the first sheet and converts it to CSV format, then delegates
    to the CSV auto-parser which handles edge lists, adjacency matrices,
    and adjacency lists.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError(
            "Excel support requires openpyxl. Install with: pip install openpyxl"
        )

    data = file.read()
    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("Excel file has no active sheet")

    # Convert to CSV in memory
    lines = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(cell) if cell is not None else "" for cell in row]
        lines.append(",".join(cells))

    wb.close()

    csv_text = "\n".join(lines)
    csv_bytes = csv_text.encode("utf-8")

    return parse_csv_auto(io.BytesIO(csv_bytes), directed=directed)
